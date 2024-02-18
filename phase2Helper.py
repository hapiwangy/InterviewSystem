def copyfromtwotoone(excelname: str, fromworksheet: str, toworksheetname: str ,index:int)-> None:
    '''
        get the new data in specific and add to the new 
    '''
    from openpyxl import load_workbook
    import helper
    excelbook = load_workbook(f"{excelname}.xlsx")
    # establish sheetnames
    SheetNames = excelbook.sheetnames
    # from workbook
    from_work = excelbook.worksheets[SheetNames.index(fromworksheet)]
    # to workbook
    to_work = excelbook.worksheets[SheetNames.index(toworksheetname)]
    # tks workbook
    tks_work = excelbook.worksheets[-1]
    # adding data in from to to 
    last_row = to_work.max_row
    last_tks_row = tks_work.max_row
    headers = []
    header_to_index = {}
    pass_letter = []
    fail_letter = []
    for index, item in enumerate(from_work):
        # 標頭行不需要進行複製
        if index == 0: 
            for index, i in enumerate(item):
                header_to_index[i.value] = index
                headers.append(i.value)
        else:
            # 下面這個決定要把資料新增到哪裡
            # 這裡的數字會根據通過和職位的欄位做更改
            which = wheretogo(item[header_to_index['通過']], item[header_to_index['職位']])
            addingdatas(excelbook, SheetNames, which, item)
            pass_letter, fail_letter = turningToList(pass_letter, fail_letter, which, item, header_to_index)
            if which == "ThanksList":
                last_tks_row += 1
                for index, i in enumerate(item):
                    tks_work[f"{chr(ord('A') + index)}{last_tks_row}"] = None     
            else:
                last_row += 1
                for index, i in enumerate(item):
                    to_work[f"{chr(ord('A') + index)}{last_row}"] = i.value
                       
    # do the letter order management
    pass_dec = False
    fail_dec = False
    if pass_letter:
        helper.NameListTomail(pass_letter, "通過二階段面試名單", "phase2", "pass")
        pass_dec = not pass_dec
    if fail_letter:
        helper.NameListTomail(fail_letter, "第二階段感謝信名單", "phase2", "thanks")
        fail_dec = not fail_dec
    if not pass_dec and not fail_dec:
        print("本次沒有人通過以及被刷掉，請透過人工確認是否有誤")
    
    # remove worksheet and adding new headers
    excelbook.remove(from_work)
    new_sheet = excelbook.create_sheet(fromworksheet, index = index)
    for i, x in enumerate(headers):
        new_sheet[f"{chr(ord('A') + i)}1"] = x

    excelbook.save(f'{excelname}.xlsx')
# deciding where to go with state and title
def wheretogo(state, title)->str:
    tech = ('tai', 'rdi')
    non_tech = ('moi', 'oai', 'psi')
    if state.value.upper() == 'X':
        return "ThanksList"
    else:
        if title.value.lower() in tech:
            return "techpass1"
        elif title.value.lower() in non_tech:
            return "nontechpass1"
        else:
            print("An Error has occured! when deciding where to go!")
# adding datas in specific excel
def addingdatas(excelbook, SheetNames, whichsheet, item)-> None:
    target_book = excelbook.worksheets[SheetNames.index(whichsheet)]
    last_rows = target_book.max_row
    for index, i in enumerate(item):
        if index == len(item) - 1:
            target_book[f"{chr(ord('A') + index)}{last_rows+1}"] = None
        else:
            target_book[f"{chr(ord('A') + index)}{last_rows+1}"] = i.value
# turning element into list
def turningToList(pass_letter, fail_letter, which, item, header_to_index) -> list:
    if which == "ThanksList":
        fail_letter.append(extractnep(item, header_to_index))
    elif which == "techpass1" or which == "nontechpass1":
        pass_letter.append(extractnep(item, header_to_index))
    else:
        print("an error occur! when manage datas to list")
    return (pass_letter, fail_letter)
        
# extract name, email, phone number
def extractnep(item, header_to_index) -> tuple:
    data = [str(x) for x in (item[header_to_index['姓名']].value,item[header_to_index['郵件']].value,item[header_to_index['通過']].value)]
    return tuple(data)


    