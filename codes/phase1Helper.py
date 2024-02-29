def copyfromtwotoone(excelname: str, fromworksheet: str, toworksheetname: str)-> None:
    '''
        get the new data in specific and add to the new 
    '''
    from openpyxl import load_workbook
    import helper
    excelbook = load_workbook(f"{excelname}.xlsx")
    # establish sheetnames
    SheetNames = excelbook.sheetnames
    # from workbook
    from_work = excelbook.worksheets[SheetNames.index(fromworksheet)]
    # to workbook
    to_work = excelbook.worksheets[SheetNames.index(toworksheetname)]
    
    # adding data in from to to 
    last_row = to_work.max_row
    headers = []
    header_to_index = {}
    pass_letter = []
    fail_letter = []
    for index, item in enumerate(from_work):
        # 標頭行不需要進行複製
        if index == 0: 
            for index, i in enumerate(item):
                header_to_index[i.value] = index
                headers.append(i.value)
        else:
            last_row += 1
            # 下面這個決定要把資料新增到哪裡
            # 這裡的數字會根據通過和職位的欄位做更改
            which = wheretogo(item[header_to_index['通過']], item[header_to_index['職位']])
            addingdatas(excelbook, SheetNames, which, item)
            pass_letter, fail_letter = turningToList(pass_letter, fail_letter, which, item, header_to_index)
            for index, i in enumerate(item):
                if index == 0:
                    continue
                to_work[f"{chr(ord('A') + index - 1)}{last_row}"] = i.value
    # do the letter order management
    pass_dec = False
    fail_dec = False
    if pass_letter:
        helper.NameListTomail(pass_letter, "通過一階段面試名單", "phase1", "pass", header_to_index)
        pass_dec = not pass_dec
    if fail_letter:
        helper.NameListTomail(fail_letter, "第一階段感謝信名單", "phase1", "thanks", header_to_index)
        fail_dec = not fail_dec
    if not pass_dec and not fail_dec:
        print("本次沒有人通過以及被刷掉，請透過人工確認是否有誤")
    
    # remove worksheet and adding new headers
    excelbook.remove(from_work)
    new_sheet = excelbook.create_sheet(fromworksheet, index = 1)
    for i, x in enumerate(headers):
        new_sheet[f"{chr(ord('A') + i)}1"] = x

    excelbook.save(f'{excelname}.xlsx')
# deciding where to go with state and title
def wheretogo(state, title)->str:
    tech = ('tai', 'rdi')
    non_tech = ('moi', 'oai', 'psi')
    if state.value.upper() == 'X':
        return "ThanksList"
    else:
        if title.value.lower() in tech:
            return "techpass1"
        elif title.value.lower() in non_tech:
            return "nontechpass1"
        else:
            print("An Error has occured! when deciding where to go!")

# adding datas in specific excel
def addingdatas(excelbook, SheetNames, whichsheet, item)-> None:
    import helper
    # 更改方向: 根據要加到哪個workbook裡面，再決定要透過什麼樣的方是來做添加
    target_book = excelbook.worksheets[SheetNames.index(whichsheet)]
    last_rows = target_book.max_row
    currentdate = helper.ReturnCurrentDate()
    if whichsheet == 'ThanksList':
        # 計算時間戳記並增加到最前面
        target_book[f"B{last_rows+1}"] = currentdate    
        # 把除了通過以外的東西都加到新的sheet裡面
        for index, i in enumerate(item):
            # 不要管通過的部分
            if index == 0:
                continue
            else:
                target_book[f"{chr(ord('A') + index + 1)}{last_rows+1}"] = i.value    
    elif whichsheet == 'techpass1' or whichsheet == 'nontechpass1':
        target_book[f"B{last_rows+1}"] = currentdate  
        for index, i in enumerate(item):
            if index == 0:
                target_book[f"{chr(ord('A') + index + 1)}{last_rows+1}"] = None
            else:
                target_book[f"{chr(ord('A') + index + 1)}{last_rows+1}"] = i.value
    else:
        print('沒有這個選項喔 ==')
# turning element into list
def turningToList(pass_letter, fail_letter, which, item, header_to_index) -> list:
    if which == "ThanksList":
        fail_letter.append(item)
    elif which == "techpass1" or which == "nontechpass1":
        pass_letter.append(item)
    else:
        print("an error occur! when manage datas to list")
    return (pass_letter, fail_letter)
        
# extract name, email, phone number
def extractnep(item, header_to_index) -> tuple:
    titles = list(header_to_index.keys())
    specdata = [item[header_to_index[x]].value for x in titles] 
    data = [str(x) for x in specdata]
    return tuple(data)

# making the dictionary about column to index
def worksheet_column_index(excelbook, SheetNames, worksheet)->dict:
    temp_dict = {}
    # 找到每一個第一排的東西
    current_work_sheet = excelbook.worksheets[SheetNames.index(worksheet)]
    current_work_sheet = list(current_work_sheet)[0]
    for index, i in enumerate(current_work_sheet):
        temp_dict[i.value] = index  
    return temp_dict



    