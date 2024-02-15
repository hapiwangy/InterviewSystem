def copyfromtwotoone(excelname: str, fromworksheet: str, toworksheetname: str)-> None:
    '''
        get the new data in specific and add to the new 
    '''
    from openpyxl import load_workbook
    excelbook = load_workbook(f"{excelname}.xlsx")
    # establish sheetnames
    SheetNames = excelbook.sheetnames
    # from workbook
    from_work = excelbook.worksheets[SheetNames.index(fromworksheet)]
    # to workbook
    to_work = excelbook.worksheets[SheetNames.index(toworksheetname)]
    
    # adding data in from to to 
    last_row = to_work.max_row
    headers = []
    for index, item in enumerate(from_work):
        # 標頭行不需要進行複製
        if index == 0: 
            for i in item:
                headers.append(i.value)
        else:
            last_row += 1
            for index, i in enumerate(item):
                to_work[f"{chr(ord('A') + index)}{last_row}"] = i.value
    # remove worksheet and adding new headers
    excelbook.remove(from_work)
    new_sheet = excelbook.create_sheet(fromworksheet, index = 1)
    for i, x in enumerate(headers):
        new_sheet[f"{chr(ord('A') + i)}1"] = x

    excelbook.save(f'{excelname}.xlsx')


